from pprint import pprint
from PureCloudPlatformClientV2.rest import ApiException
import os
import time
import openpyxl as xl

def get_all_user_prompt_pg_count(api_instance, o_logger):
    try:
        api_response = api_instance.get_architect_prompts()
        # pprint(api_response)

        total_result_pages = api_response.page_count
        log_msg = 'User Prompt ID Page count: ' + str(total_result_pages)
        o_logger.info(log_msg)
        return total_result_pages
    except ApiException as e:
        print("Exception calling get_architect_prompts: %s\n" % e)


def retrieve_user_prompt_names(pg_count, api_instance, o_logger):
    # page_number = 1  # int | Page number (optional) (default to 1)
    # page_size = 25  # int | Page size (optional) (default to 25)
    # name = ['name_example'] # list[str] | Name (optional)
    # description = 'description_example' # str | Description (optional)
    # name_or_description = 'name_or_description_example' # str | Name or description (optional)
    # sort_by = 'id' # str | Sort by (optional) (default to 'id')
    # sort_order = 'asc' # str | Sort order (optional) (default to 'asc')
    up_name_list = []

    log_msg = 'Listing found user prompts'
    o_logger.info(log_msg)

    pg_index = 1
    while pg_index <= pg_count:
        try:
            log_msg = 'Reading page ' + str(pg_index)
            o_logger.info(log_msg)

            # Get a pageable list of user prompts
            api_response = api_instance.get_architect_prompts(page_number=pg_index)
            # pprint(api_response)

            user_prompt_id_list = api_response.entities
            # pprint(user_prompt_id_list)

            for item in user_prompt_id_list:
                up_id = item.id
                up_name = item.name
                # print(f'{up_id}:{up_name}')
                log_msg = 'User Prompt ID=' + up_id + ":User Prompt Name=" + up_name
                o_logger.info(log_msg)
                up_name_list.append(up_name)

        except ApiException as e:
            print("Exception when calling GetArchitectPromptsRequest->get_architect_prompts: %s\n" % e)

        pg_index += 1

    return up_name_list


def get_all_up_dependencies_pg_count(api_instance, up_name, flow_type, o_logger):
    try:
        print(f'Processing User Prompt Name: {up_name}')
        name = up_name
        object_type = 'USERPROMPT'
        consuming_resources = True
        consuming_resource_type = flow_type

        try:
            # Get Dependency Tracking objects that have a given display name
            api_response = api_instance.get_architect_dependencytracking(name,
                                                                         object_type=object_type,
                                                                         consuming_resources=consuming_resources,
                                                                         consuming_resource_type=consuming_resource_type
                                                                         )
            # pprint(api_response)
            total_result_pages = api_response.page_count
            for item in api_response.entities:
                if len(item.consuming_resources) == 0:
                    total_result_pages = 0
                    print('No dependencies found.')
            log_msg = 'Dependency Page count: ' + str(total_result_pages)
            o_logger.info(log_msg)
            return total_result_pages
        except ApiException as e:
            print("Exception calling get_architect_dependencytracking : %s\n" % e)

    except ApiException as e:
        print("Exception executing get_all_up_dependencies_pg_count : %s\n" % e)


def retrieve_dependencies(output_wb, output_xlsx_sheet, pg_count, api_instance, up_name, flow_type, o_logger):
    colmapping = {
        "prompt_name": 1,
        "flow_type": 2,
        "flow_name":3,
        "flow_version":4
    }

    try:
        print(f'Retrieving dependencies for User Prompt Name: {up_name} used by {flow_type}')
        name = up_name
        object_type = 'USERPROMPT'
        consuming_resources = True
        consuming_resource_type = flow_type

        log_msg = 'Listing dependencies for User Prompt ' + up_name + ":"
        o_logger.info(log_msg)

        pg_index = 1
        while pg_index <= pg_count:
            try:
                log_msg = 'Reading page ' + str(pg_index)
                o_logger.info(log_msg)
                page_number = pg_index

                # Get a specific page result
                api_response = \
                    api_instance.get_architect_dependencytracking(
                        name,
                        object_type=object_type,
                        consuming_resources=consuming_resources,
                        consuming_resource_type=consuming_resource_type
                        )
                # pprint(api_response)

                dependency_flow_list = api_response.entities
                # pprint(dependency_flow_list)

                next_blank_row = output_xlsx_sheet.max_row + 1
                # print(next_blank_row)

                for item in dependency_flow_list:
                    for critem in item.consuming_resources:
                        dpf_id = critem.id
                        dpf_name = critem.name
                        dpf_type = critem.type
                        dpf_ver = critem.version

                        # Write to output spreadsheet
                        output_xlsx_sheet.cell(next_blank_row, 1).value = up_name
                        output_xlsx_sheet.cell(next_blank_row, 2).value = dpf_type
                        output_xlsx_sheet.cell(next_blank_row, 3).value = dpf_name
                        output_xlsx_sheet.cell(next_blank_row, 4).value = dpf_ver
                        next_blank_row += 1
                        print(next_blank_row)
                        # print(f'{dpf_id}:{dpf_name}:{dpf_type}:{dpf_ver}')
                        log_msg = ">>>" + dpf_id + ":" + dpf_name + ":" + dpf_type + ":" + dpf_ver
                        o_logger.info(log_msg)

                pg_index += 1

            except ApiException as e:
                print("Exception when calling get_architect_dependencytracking_consumingresources: %s\n" % e)

    except ApiException as e:
        print("Exception executing retrieve_dependencies: %s\n" % e)

    return output_wb, output_xlsx_sheet


def read_flow_types(filepath):
    wb = xl.load_workbook(filepath)
    sheet = wb['Sheet1']
    totalrows = sheet.max_row
    # print(f'The {filepath} has {totalrows} rows.') for debugging only.

    # print('The following flow types are read.') for debugging only
    flow_type_list =[]

    for row in range(1, totalrows + 1):
        flowtype = sheet.cell(row, 1)
        # print(flowtype.value) for debugging only.
        flow_type_list.append(flowtype.value)

    return flow_type_list


def prepare_output_xlsx(output_template_filename):

    wb = xl.load_workbook(output_template_filename)
    sheet = wb['Sheet1']
    return wb, sheet

def write_output_to_xlsx(wb, sheet, curr_row, curr_col, output_cell_value):
    sheet.cell(curr_row, curr_col).value = output_cell_value