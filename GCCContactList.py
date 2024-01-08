# This is a core GCC library for manipulating contact list
import PureCloudPlatformClientV2
from PureCloudPlatformClientV2.rest import ApiException
from pprint import pprint
import os
import openpyxl as xl
import pandas as pd


def fix_exported_csv_filename(exptcsvfilename):
        removeindex = exptcsvfilename.rfind('.')
        return exptcsvfilename[0:removeindex]


def get_contactlistid_from_filename(csvfilename):
    removeindex = csvfilename.rfind('-')
    return csvfilename[0:removeindex]


def convert_csv_as_xlsx(dirname, csvfilename):
    # Load your spreadsheet as wb object.
    csvpath = dirname + csvfilename + '.csv'
    xlsxpath = dirname + csvfilename + '.xlsx'

    df = pd.read_csv(csvpath)
    df.to_excel(xlsxpath)

    #  list_dir_content(dirname)  For debugging only.


def get_contactids_from_contact_list(dirname, csvfilename, safelist, o_logger):
    contactlist = dirname + csvfilename + '.xlsx'

    colmapping = {
        "record_number": 1,
        "inin-outbound-id": 2,
        "phone": 3,
        "name": 4,
        "delay": 5,
        "callbackId": 6,
        "ContactCallable": 7,
        "ZipCodeAutomaticTimeZone": 8,
        "CallRecordLastAttempt-phone": 9,
        "CallRecordLastResult-phone": 10,
        "CallRecordLastAgentWrapup-phone": 11,
        "SmsLastAttempt-phone": 12,
        "SmsLastResult-phone": 13,
        "Callable-phone": 14,
        "AutomaticTimeZone-phone": 15
    }

    wb = xl.load_workbook(contactlist)
    sheet = wb['Sheet1']
    totalrows = sheet.max_row

    log_msg = 'The ' + contactlist + ' has ' + str(totalrows) + ' rows.'
    o_logger.info(log_msg)

    # The discounted prices go righ of the price column.
    contactidcol = int(colmapping.get("inin-outbound-id"))
    recordstatecol = int(colmapping.get("CallRecordLastResult-phone"))

    log_msg = 'The following ids will be deleted:'
    o_logger.info(log_msg)

    myidlist =[]
    selectedrecordcounter = 0

    for row in range(2,totalrows + 1):
        contactid = sheet.cell(row, contactidcol)
        recstatus = sheet.cell(row, recordstatecol)

        if recstatus.value in safelist:
            next_contact_id = (contactid.value)
            o_logger.info(next_contact_id)
            myidlist.append(next_contact_id)
            selectedrecordcounter += 1
            # Below line prevents overloading API by limiting contact id to 1000.
            if selectedrecordcounter > 1000:
                break

    return myidlist


def remove_contacts_from_contactlist(mygcp, mytoken, contact_list_id, contact_ids, o_logger):
    id_count = len(contact_ids)

    # instantiate UsersApi class as api_instance and then call the post_users API.
    api_instance = mygcp.OutboundApi(mytoken)

    try:
        # Delete contacts from a contact list.
        api_instance.delete_outbound_contactlist_contacts(contact_list_id, contact_ids)
        log_msg = str(id_count) + ' records removed.'
        o_logger.info(log_msg)
    except ApiException as e:
        log_msg = ("Exception occurred with API call" % e)


