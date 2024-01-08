import openpyxl as xl


def read_SafeList(slpath):
    wb = xl.load_workbook(slpath)
    sheet = wb['Sheet1']
    totalrows = sheet.max_row
    # print(f'The {slpath} has {totalrows} rows.') for debugging only.

    # print('The following wrap up codes are read.') for debugging only
    mywclist =[]

    for row in range(1, totalrows + 1):
        wrapupcode = sheet.cell(row, 1)
        # print(wrapupcode.value) for debugging only.
        mywclist.append(wrapupcode.value)

    return mywclist


